# -*- coding: utf-8 -*-

from __future__ import absolute_import, division, print_function, unicode_literals
import datetime as dt
try:
    from Tkinter import Tk
except ImportError:
    from tkinter import Tk
try:
    import tkFileDialog
except ImportError:
    from tkinter import filedialog as tkFileDialog
import openpyxl as op
import argparse
import os.path
import sys
import re
import csv
from . import rfc4180

__metaclass__ = type

def parseargs():
    pa = argparse.ArgumentParser(description=
            'Exports multiple CSV files from an Excel *.xlsx Workbook')
    pa.add_argument('-f', metavar='EXCELFILE',
                    help='The Excel file to export. ' +
                    'If omitted, a graphical file chooser will be used.')
    pa.add_argument('-o', metavar='OUTPUTDIRECTORY',
                    help='The output directory. Default is the current ' +
                    'directory if EXCELFILE was given, otherwise a ' +
                    'file chooser will be used as well.')
    args = pa.parse_args(sys.argv[1:])
    return vars(args)

def _stringify(dat):
    if not isinstance(dat, basestring):
        return str(dat).encode('utf-8')
    else:
        return dat.encode('utf-8')

def _transmap(dat):
    transmap = {
        # empty cells are going to be empty strings
        None: '',
        # workaround for bug in openpyxl
        # https://bitbucket.org/openpyxl/openpyxl/issues/674/ 
        dt.datetime(1899, 12, 30, 0, 0): dt.time(0, 0),
        dt.datetime(1899, 12, 31, 0, 0): dt.datetime(1900, 1, 1, 0, 0),
    }
    return transmap[dat] if dat in transmap else dat

def _datefix(dat):
    # if typ is datetime.datetime and time-part is 0:0:0,
    # covert to datetime.date (assume xlsx cell-type is "Date").
    if (type(dat) == dt.datetime and
        (dat.hour, dat.minute, dat.second) == (0, 0, 0)):
        dat = dat.date()
    return dat

def transform(l):
    l = [_transmap(f) for f in l]
    l = [_datefix(f) for f in l]
    l = [_stringify(f) for f in l]
    return l
    
def write_csv(data, outfile):
    with open(outfile, 'wb') as fout:
        writer = csv.writer(fout, dialect='RFC4180')
        writer.writerows(data)

def main(xlsxfile, out_dir, sheets=None):
    rfc4180.RFC4180()
    out_prefix = os.path.splitext(os.path.basename(xlsxfile))[0]
    wb = op.load_workbook(xlsxfile, data_only=True)
    sheetnames = sheets or wb.sheetnames
    for sn in sheetnames:
        outfile = os.path.join(out_dir, out_prefix + '_' +
                               re.sub(r'\s+', '_', sn) + '.csv')
        data = []
        try:
            sheet = wb.get_sheet_by_name(sn)
        except KeyError as e:
            print(e.message)
            continue
        for l in sheet.values:
            data.append(transform(l))
        write_csv(data, outfile)

def main_gui():
    home = os.path.expanduser('~')
    xlsxfile = parseargs()['f']
    out_dir = parseargs()['o']
    if xlsxfile is None:
        root = Tk()
        root.withdraw()
        f = tkFileDialog.askopenfile(title='Choose file to convert',
                                      filetypes=[('xlsx', '*.xlsx')],
                                      initialdir=home)
        if f:
            xlsxfile = f.name
            f.close()
        else:
            sys.exit()
        if out_dir is None:
            out_dir = tkFileDialog.askdirectory(title='Choose output directory',
                                                initialdir=home)
            if not out_dir:
                sys.exit()
        root.destroy()
    if not out_dir:
        out_dir = os.getcwd()
    main(xlsxfile, out_dir)
    
if __name__ == '__main__':
    main_gui()
